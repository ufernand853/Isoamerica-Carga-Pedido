import os
from datetime import datetime
import argparse
import importlib.util

if importlib.util.find_spec("tkinter") is not None:
    import tkinter as tk
    from tkinter import filedialog
    TK_AVAILABLE = True
else:
    tk = None
    filedialog = None
    TK_AVAILABLE = False

import pandas as pd
from openpyxl import load_workbook

# === CONFIGURACIÓN ===
CONFIG_ARCHIVOS = "config_archivos.txt"
PEDIDO_FILE = "Planilla pedido 10.12.2025 Destino.xlsx"
LISTADO_FILE = "Listado general para PLANILLAS TRADU BRs.xlsx"
OUTPUT_FILE = "Planilla pedido 10.12.2025 Destino_COMPLETADA.xlsx"
# Guarda la salida configurada explícitamente en config_archivos.txt (si existe)
_OUTPUT_FILE_CONFIGURADO = None


def _leer_config_archivos(path=CONFIG_ARCHIVOS):
    """
    Carga nombres de archivos desde un archivo de texto opcional con formato clave=valor.
    Claves reconocidas: pedido, listado, salida/output.
    """

    if not os.path.isfile(path):
        return {}

    config = {}
    with open(path, "r", encoding="utf-8") as f:
        for linea in f:
            entrada = linea.strip()
            if not entrada or entrada.startswith("#") or "=" not in entrada:
                continue

            clave, valor = entrada.split("=", 1)
            clave = clave.strip().lower()
            valor = valor.strip()
            if valor:
                config[clave] = valor
    return config


def _aplicar_config_archivos(config):
    global PEDIDO_FILE, LISTADO_FILE, OUTPUT_FILE, _OUTPUT_FILE_CONFIGURADO

    PEDIDO_FILE = config.get("pedido", PEDIDO_FILE)
    LISTADO_FILE = config.get("listado", LISTADO_FILE)
    salida_config = config.get("salida", config.get("output"))

    if salida_config:
        OUTPUT_FILE = salida_config
        _OUTPUT_FILE_CONFIGURADO = salida_config


_CONFIG_ARCHIVOS = _leer_config_archivos()
if _CONFIG_ARCHIVOS:
    _aplicar_config_archivos(_CONFIG_ARCHIVOS)


def _leer_config_archivos(path=CONFIG_ARCHIVOS):
    """
    Carga nombres de archivos desde un archivo de texto opcional con formato clave=valor.
    Claves reconocidas: pedido, listado, salida/output.
    """

    if not os.path.isfile(path):
        return {}

    config = {}
    with open(path, "r", encoding="utf-8") as f:
        for linea in f:
            entrada = linea.strip()
            if not entrada or entrada.startswith("#") or "=" not in entrada:
                continue

            clave, valor = entrada.split("=", 1)
            clave = clave.strip().lower()
            valor = valor.strip()
            if valor:
                config[clave] = valor
    return config


def _aplicar_config_archivos(config):
    global PEDIDO_FILE, LISTADO_FILE, OUTPUT_FILE

    PEDIDO_FILE = config.get("pedido", PEDIDO_FILE)
    LISTADO_FILE = config.get("listado", LISTADO_FILE)
    OUTPUT_FILE = config.get("salida", config.get("output", OUTPUT_FILE))


_CONFIG_ARCHIVOS = _leer_config_archivos()
if _CONFIG_ARCHIVOS:
    _aplicar_config_archivos(_CONFIG_ARCHIVOS)

# Nombre de hoja (None = primera)
PEDIDO_SHEET = None
LISTADO_SHEET = None

# Nombre de la columna clave en la planilla de pedido (columna E en tu ejemplo)
PEDIDO_KEY_COL = "Codigo Principal"

# Nombre de la columna clave en el listado general
LISTADO_KEY_COL = "Codigo"

# Mapeo desde columnas del LISTADO a columnas destino en la planilla de pedido
COLUMN_MAPPING = {
    # destino_en_pedido : origen_en_listado
    "EAN - Cod Barras": "EAN",                      # Columna D
    "Descrição": "Descripcion",                     # Columna F - traducción
    "Marca": "Fabricante",                          # Columna G (a falta de columna Marca, usamos Fabricante)
    "Pais de Origem": "Pais",                       # Columna H
    "NCM": "NCM",                                   # Columna I
    "Peso Neto Unitario": "Peso",                   # Columna R
    "Nome do Fabricante - Razão Social": "Fabricante",   # Columna W
    "Endereço do Fabricante - Rua - Numero - Cidade - Estado - CEP": "Ubicacion",  # Columna X
}

def cargar_listado(path, sheet_name=None):
    # Si no se especifica hoja, usamos la primera
    if sheet_name is None:
        listado = pd.read_excel(path, sheet_name=0)
    else:
        listado = pd.read_excel(path, sheet_name=sheet_name)

    # En tu archivo actual el listado tiene 9 columnas sin encabezados lógicos:
    # asumimos el orden: EAN, Codigo, Descripcion, Pais, NCM, Peso, Fabricante, Ubicacion, Extra
    if listado.shape[1] == 9:
        listado.columns = [
            "EAN",        # col 0
            "Codigo",     # col 1
            "Descripcion",# col 2
            "Pais",       # col 3
            "NCM",        # col 4
            "Peso",       # col 5
            "Fabricante", # col 6
            "Ubicacion",  # col 7
            "Extra",      # col 8 (no usada)
        ]
    return listado

def construir_diccionario(listado, key_col):
    """
    Construye un índice: clave -> fila (Series)
    para búsqueda rápida por código.
    """
    listado_key = listado.copy()
    listado_key[key_col] = listado_key[key_col].astype(str).str.strip()
    listado_key = listado_key.drop_duplicates(subset=key_col, keep="first")
    return listado_key.set_index(key_col)

def completar_planilla_pedido(pedido_path, listado_path, output_path):
    # --- Cargar listado general ---
    listado = cargar_listado(listado_path, LISTADO_SHEET)
    indexado = construir_diccionario(listado, LISTADO_KEY_COL)

    # --- Cargar planilla de pedido como DataFrame para identificar columnas ---
    # header=4 porque los encabezados reales están en la fila 5 de Excel
    if PEDIDO_SHEET is None:
        pedido_raw = pd.read_excel(pedido_path, sheet_name=0, header=4)
    else:
        pedido_raw = pd.read_excel(pedido_path, sheet_name=PEDIDO_SHEET, header=4)

    # La primera fila de pedido_raw contiene los nombres de las columnas
    header_row = pedido_raw.iloc[0]
    pedido_data = pedido_raw[1:].copy()
    pedido_data.columns = header_row

    # Normalizar la columna clave de pedido
    pedido_data[PEDIDO_KEY_COL] = pedido_data[PEDIDO_KEY_COL].astype(str).str.strip()

    # --- Abrir el Excel original con openpyxl para mantener formato ---
    wb = load_workbook(pedido_path)
    ws = wb[wb.sheetnames[0]] if PEDIDO_SHEET is None else wb[PEDIDO_SHEET]

    # La primera fila de datos (Item 1) es la fila 6 de Excel.
    # En el DataFrame, la primera fila de datos tiene índice 1, por lo que:
    # fila_excel = índice_df + 5
    total_rows = 0
    matched_rows = 0

    for idx, row in pedido_data.iterrows():
        codigo = str(row.get(PEDIDO_KEY_COL, "")).strip()
        if not codigo or codigo.lower() == "nan":
            continue

        total_rows += 1

        if codigo in indexado.index:
            matched_rows += 1
            fuente = indexado.loc[codigo]
            excel_row = idx + 5  # ver comentario arriba

            for destino_col, origen_col in COLUMN_MAPPING.items():
                if origen_col not in fuente.index:
                    continue

                valor = fuente[origen_col]

                if destino_col not in pedido_data.columns:
                    continue

                col_idx = list(pedido_data.columns).index(destino_col)
                excel_col = col_idx + 1  # A=1, B=2, etc.

                ws.cell(row=excel_row, column=excel_col, value=valor)

    wb.save(output_path)

    print(f"Filas procesadas (con código en la columna E): {total_rows}")
    print(f"Filas con coincidencia en el listado: {matched_rows}")
    print(f"Archivo generado: {output_path}")


def _generar_output(pedido_path, explicit_output=None):
    """Devuelve la ruta de salida priorizando la configuración del TXT."""

    if explicit_output:
        return explicit_output

    # Si el usuario definió una salida en config_archivos.txt, úsala siempre
    if _OUTPUT_FILE_CONFIGURADO:
        return _OUTPUT_FILE_CONFIGURADO

    # Si no hay configuración específica, sugerimos una salida junto al pedido
    folder = os.path.dirname(pedido_path) or "."
    nombre_archivo = os.path.basename(pedido_path)
    nombre, ext = os.path.splitext(nombre_archivo)
    fecha = datetime.now().strftime("%Y-%m-%d")
    if not ext:
        ext = ".xlsx"

    return os.path.join(folder, f"{nombre}_procesada_{fecha}{ext}")


def _solicitar_ruta(mensaje, predeterminada=None):
    ruta = input(mensaje).strip()
    if not ruta and predeterminada:
        return predeterminada
    return ruta or None


def _ruta_predeterminada(ruta):
    return ruta if os.path.exists(ruta) else None


def _seleccionar_archivo_gui(titulo, archivo_sugerido):
    return filedialog.askopenfilename(
        title=titulo,
        initialdir=os.path.dirname(archivo_sugerido) or ".",
        initialfile=os.path.basename(archivo_sugerido),
        filetypes=[("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
    )


def _seleccionar_salida_gui(titulo, ruta_sugerida):
    carpeta = os.path.dirname(ruta_sugerida) or "."
    archivo = os.path.basename(ruta_sugerida) or os.path.basename(OUTPUT_FILE)
    return filedialog.asksaveasfilename(
        title=titulo,
        defaultextension=".xlsx",
        initialdir=carpeta,
        initialfile=archivo,
        filetypes=[("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
    )


def seleccionar_archivos_cli(args):
    pedido_path = args.pedido or _ruta_predeterminada(PEDIDO_FILE)
    listado_path = args.listado or _ruta_predeterminada(LISTADO_FILE)

    if not pedido_path:
        pedido_path = _solicitar_ruta(
            f"Ruta de la planilla de pedido [{PEDIDO_FILE}]: ", predeterminada=PEDIDO_FILE
        )

    if not listado_path:
        listado_path = _solicitar_ruta(
            f"Ruta del listado general [{LISTADO_FILE}]: ", predeterminada=LISTADO_FILE
        )

    output_path = _generar_output(pedido_path, args.output) if pedido_path else args.output

    if not output_path:
        output_path = _solicitar_ruta(
            f"Ruta de salida [{OUTPUT_FILE}]: ", predeterminada=OUTPUT_FILE
        )

    return pedido_path, listado_path, output_path


def seleccionar_archivos_gui(args):
    if not TK_AVAILABLE:
        raise ImportError(
            "Tkinter no está disponible en este entorno. Instálalo (ej. python3-tk) o usa --cli."
        )

    root = tk.Tk()
    root.withdraw()

    pedido_predeterminado = args.pedido or _ruta_predeterminada(PEDIDO_FILE) or PEDIDO_FILE
    pedido_path = args.pedido or _seleccionar_archivo_gui(
        "Selecciona la planilla de pedido", pedido_predeterminado
    )

    listado_predeterminado = args.listado or _ruta_predeterminada(LISTADO_FILE) or LISTADO_FILE
    listado_path = args.listado or _seleccionar_archivo_gui(
        "Selecciona el listado general", listado_predeterminado
    )

    output_sugerida = _generar_output(pedido_path, args.output) if pedido_path else args.output
    output_sugerida = output_sugerida or OUTPUT_FILE
    output_path = _seleccionar_salida_gui(
        "Guardar planilla completada como", output_sugerida
    ) if output_sugerida else None

    root.destroy()
    return pedido_path or None, listado_path or None, output_path or None


def build_parser():
    parser = argparse.ArgumentParser(
        description=(
            "Completa la planilla de pedido usando datos del listado general. "
            "Si no se especifican rutas, se solicitarán por consola."
        )
    )
    parser.add_argument("--pedido", help="Ruta del archivo de planilla de pedido")
    parser.add_argument("--listado", help="Ruta del archivo con el listado general")
    parser.add_argument(
        "--output",
        help=(
            "Ruta del archivo de salida. Si no se especifica, se generará junto "
            "a la planilla de pedido con sufijo procesada_FECHA."
        ),
    )
    parser.add_argument(
        "--pedido-sheet",
        dest="pedido_sheet",
        default=PEDIDO_SHEET,
        help="Nombre de la hoja de pedido (por defecto, la primera)",
    )
    parser.add_argument(
        "--listado-sheet",
        dest="listado_sheet",
        default=LISTADO_SHEET,
        help="Nombre de la hoja del listado (por defecto, la primera)",
    )
    parser.add_argument(
        "--cli",
        action="store_true",
        help="Usar la selección de archivos por consola en lugar de la interfaz Tkinter.",
    )
    return parser

if __name__ == "__main__":
    parser = build_parser()
    args = parser.parse_args()

    # Permitir sobrescribir hojas mediante argumentos
    if args.pedido_sheet is not None:
        PEDIDO_SHEET = args.pedido_sheet
    if args.listado_sheet is not None:
        LISTADO_SHEET = args.listado_sheet

    if args.cli:
        pedido_path, listado_path, output_path = seleccionar_archivos_cli(args)
    else:
        if TK_AVAILABLE:
            pedido_path, listado_path, output_path = seleccionar_archivos_gui(args)
        else:
            print(
                "Tkinter no está instalado en este intérprete. "
                "Usando el modo consola (--cli)."
            )
            pedido_path, listado_path, output_path = seleccionar_archivos_cli(args)

    if pedido_path and listado_path and output_path:
        completar_planilla_pedido(pedido_path, listado_path, output_path)
    else:
        print("Selección cancelada. No se procesó ninguna planilla.")
